from __future__ import annotations

import base64
import csv
import io
import json
from dataclasses import dataclass
from datetime import date, datetime

from .schemas import FilePayload


IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "webp"}
IMAGE_MIME_TYPES = {
    "image/png",
    "image/jpeg",
    "image/jpg",
    "image/webp",
}
AUDIO_EXTENSIONS = {"mp3", "wav", "m4a", "ogg", "webm", "mpeg"}
AUDIO_MIME_TYPES = {
    "audio/mpeg",
    "audio/mp3",
    "audio/wav",
    "audio/x-wav",
    "audio/mp4",
    "audio/m4a",
    "audio/ogg",
    "audio/webm",
}


@dataclass(slots=True)
class ParsedFile:
    text: str | None
    mode: str | None
    supported: bool
    message: str | None = None
    action_hint: str | None = None
    structured_records: list[dict] | None = None


def parse_attachment(file_payload: FilePayload | None, max_chars: int = 8000) -> ParsedFile:
    if file_payload is None:
        return ParsedFile(text=None, mode=None, supported=True)

    try:
        raw = base64.b64decode(file_payload.base64)
    except Exception:
        return ParsedFile(
            text=None,
            mode=None,
            supported=False,
            message="Nao foi possivel decodificar o anexo recebido.",
        )

    extension = (file_payload.extensao or "").lower()
    mime_type = (file_payload.mime_type or "").lower()

    if extension in {"txt", "md", "csv", "json"} or mime_type in {
        "text/plain",
        "text/csv",
        "application/json",
    }:
        if extension == "csv" or mime_type == "text/csv":
            text, structured_records, action_hint = _extract_csv(raw)
            return ParsedFile(
                text=_truncate(text or _decode_text(raw), max_chars),
                mode="csv",
                supported=True,
                action_hint=action_hint,
                structured_records=structured_records,
            )

        text = _decode_text(raw)
        if extension == "json" or mime_type == "application/json":
            text = _pretty_json(text)
        return ParsedFile(text=_truncate(text, max_chars), mode="text", supported=True)

    if extension in {"pdf"} or mime_type == "application/pdf":
        text = _extract_pdf(raw)
        if text:
            return ParsedFile(text=_truncate(text, max_chars), mode="pdf", supported=True)
        return ParsedFile(
            text=None,
            mode="pdf",
            supported=False,
            message="Recebi o PDF, mas a extracao de texto foi parcial ou falhou. Vou continuar com o contexto disponivel e pedir apenas o que faltar.",
        )

    if extension in {"xlsx", "xlsm"} or mime_type in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroenabled.12",
    }:
        text = _extract_xlsx(raw)
        if text:
            structured_records, action_hint = _extract_fatura_candidates(text)
            return ParsedFile(
                text=_truncate(text, max_chars),
                mode="xlsx",
                supported=True,
                action_hint=action_hint,
                structured_records=structured_records,
            )
        return ParsedFile(
            text=None,
            mode="xlsx",
            supported=False,
            message="Recebi a planilha, mas a extracao do conteudo foi parcial ou falhou. Vou continuar com o contexto disponivel e pedir apenas o que faltar.",
        )

    if extension in {"xls"} or mime_type == "application/vnd.ms-excel":
        text = _extract_xls(raw)
        if text:
            structured_records, action_hint = _extract_fatura_candidates(text)
            return ParsedFile(
                text=_truncate(text, max_chars),
                mode="xls",
                supported=True,
                action_hint=action_hint,
                structured_records=structured_records,
            )
        return ParsedFile(
            text=None,
            mode="xls",
            supported=False,
            message="Recebi a planilha XLS, mas a extracao do conteudo foi parcial ou falhou. Vou continuar com o contexto disponivel e pedir apenas o que faltar.",
        )

    if is_image_attachment(extension, mime_type):
        return ParsedFile(
            text=None,
            mode="image",
            supported=True,
            message="Recebi a imagem e vou extrair o conteudo antes de continuar a analise.",
        )

    if is_audio_attachment(extension, mime_type):
        return ParsedFile(
            text=None,
            mode="audio",
            supported=True,
            message="Recebi o audio e vou transcrever o conteudo antes de continuar a analise.",
        )

    return ParsedFile(
        text=None,
        mode=extension or mime_type or "unknown",
        supported=False,
        message="Recebi o anexo. Mesmo sem extracao completa, vou tentar identificar a acao desejada e pedir apenas os dados essenciais para concluir o processamento.",
    )


def is_image_attachment(extension: str | None, mime_type: str | None) -> bool:
    normalized_extension = (extension or "").lower()
    normalized_mime_type = (mime_type or "").lower()
    return normalized_extension in IMAGE_EXTENSIONS or normalized_mime_type in IMAGE_MIME_TYPES


def is_audio_attachment(extension: str | None, mime_type: str | None) -> bool:
    normalized_extension = (extension or "").lower()
    normalized_mime_type = (mime_type or "").lower()
    return normalized_extension in AUDIO_EXTENSIONS or normalized_mime_type in AUDIO_MIME_TYPES


def _decode_text(raw: bytes) -> str:
    for encoding in ("utf-8", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="ignore")


def _pretty_json(text: str) -> str:
    try:
        parsed = json.loads(text)
        return json.dumps(parsed, ensure_ascii=False, indent=2)
    except Exception:
        return text


def _extract_pdf(raw: bytes) -> str | None:
    try:
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(raw))
        parts: list[str] = []
        for page in reader.pages:
            parts.append(page.extract_text() or "")
        return "\n".join(part for part in parts if part.strip()) or None
    except Exception:
        return None


def _extract_xlsx(raw: bytes) -> str | None:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.extend(
                _render_sheet_rows(
                    title=sheet.title,
                    rows=sheet.iter_rows(values_only=True),
                )
            )
        return "\n".join(lines).strip() or None
    except Exception:
        return None


def _extract_xls(raw: bytes) -> str | None:
    try:
        import xlrd

        workbook = xlrd.open_workbook(file_contents=raw)
        lines: list[str] = []
        for sheet in workbook.sheets():
            lines.extend(
                _render_sheet_rows(
                    title=sheet.name,
                    rows=(sheet.row_values(row_index) for row_index in range(sheet.nrows)),
                )
            )
        return "\n".join(lines).strip() or None
    except Exception:
        return None


def _extract_csv(raw: bytes) -> tuple[str | None, list[dict] | None, str | None]:
    text = _decode_text(raw)
    if not text.strip():
        return None, None, None

    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        return text, None, None

    fieldnames = [str(field or "").strip().lstrip("\ufeff") for field in reader.fieldnames]
    lines = ["[CSV]", "Cabecalhos: " + " | ".join(fieldnames)]
    raw_rows: list[dict[str, str]] = []

    for index, row in enumerate(reader, start=1):
        normalized = {
            fieldnames[position]: _stringify_cell(row.get(original_field))
            for position, original_field in enumerate(reader.fieldnames)
            if _stringify_cell(row.get(original_field))
        }
        if normalized:
            raw_rows.append(normalized)
            lines.append(f"Registro {index}: " + json.dumps(normalized, ensure_ascii=False))

    lines.append(f"Total de registros: {max(len(lines) - 2, 0)}")
    structured_records, action_hint = _extract_client_candidates_from_csv_rows(raw_rows)
    return "\n".join(lines), structured_records, action_hint


def _extract_client_candidates_from_csv_rows(rows: list[dict[str, str]]) -> tuple[list[dict] | None, str | None]:
    if not rows:
        return None, None

    normalized_rows = [_normalize_row_keys(row) for row in rows]
    has_razao_social = any("razao_social" in row for row in normalized_rows)
    has_cnpj = any("cnpj" in row for row in normalized_rows)

    if not (has_razao_social and has_cnpj):
        return None, None

    records: list[dict] = []
    for row in normalized_rows:
        record = {
            "razao_social": row.get("razao_social"),
            "nome_fantasia": row.get("nome_fantasia"),
            "cnpj": row.get("cnpj"),
            "email": row.get("email"),
            "telefone": row.get("telefone"),
            "celular": row.get("celular"),
            "cep": row.get("cep"),
            "logradouro": row.get("logradouro") or row.get("endereco"),
            "numero": row.get("numero"),
            "bairro": row.get("bairro"),
            "cidade": row.get("cidade"),
            "uf": row.get("uf"),
        }
        records.append({key: value for key, value in record.items() if value not in (None, "")})

    return records, "sincronizar_clientes"


def _extract_fatura_candidates(text: str) -> tuple[list[dict] | None, str | None]:
    if "Empresa:" not in text or "CNPJ:" not in text or "Produto/Serviço" not in text:
        return None, None

    empresa = None
    cnpj = None
    data_vencimento = None
    periodo_referencia = None
    itens: list[dict] = []
    funcionarios: list[dict] = []
    exames: list[dict] = []
    unidade = None
    numero_funcionarios = None
    secao = "titulos"

    for line in text.splitlines():
        stripped = line.strip()
        if stripped.startswith("Empresa:"):
            empresa = stripped.split(":", 1)[1].strip()
            continue
        if stripped.startswith("CNPJ:"):
            cnpj = stripped.split(":", 1)[1].strip()
            continue
        if stripped == "Títulos":
            secao = "titulos"
            continue
        if stripped.startswith("Funcionários e Exames da Unidade:"):
            unidade = stripped.split(":", 1)[1].strip() or None
            secao = "funcionarios"
            continue
        if stripped == "Exames":
            secao = "exames"
            continue
        if stripped.startswith("Número de Funcionários:"):
            numero_funcionarios = _to_int(stripped.split(":", 1)[1].strip())
            continue
        if not stripped.startswith("Registro: "):
            continue

        try:
            data = json.loads(stripped.removeprefix("Registro: ").strip())
        except json.JSONDecodeError:
            continue

        if secao == "titulos":
            item = _parse_fatura_billing_item(data)
            if item:
                if data_vencimento is None:
                    data_vencimento = _normalize_date_value(data.get("Data Cobrança"))
                if periodo_referencia is None and data_vencimento:
                    periodo_referencia = data_vencimento[:7]
                itens.append(item)
            continue

        if secao == "funcionarios":
            funcionario = _parse_fatura_employee_row(data)
            if funcionario:
                funcionarios.append(funcionario)
            continue

        if secao == "exames":
            exame = _parse_fatura_exam_row(data)
            if exame:
                exames.append(exame)

    if not itens or not (empresa or cnpj):
        return None, None

    observacoes = _build_fatura_attachment_summary(
        unidade=unidade,
        itens=itens,
        funcionarios=funcionarios,
        exames=exames,
        numero_funcionarios=numero_funcionarios,
    )

    metadata = {
        "origem_importacao": "planilha_fatura",
        "unidade": unidade,
        "numero_funcionarios": numero_funcionarios or (len(funcionarios) if funcionarios else None),
        "funcionarios": funcionarios[:50],
        "exames": exames[:50],
        "quantidade_funcionarios_registrados": len(funcionarios),
        "quantidade_exames_registrados": len(exames),
    }

    metadata = {key: value for key, value in metadata.items() if value not in (None, "", [])}
    record = {
        "cliente": empresa,
        "cliente_cnpj": cnpj,
        "data_vencimento": data_vencimento,
        "periodo_referencia": periodo_referencia,
        "observacoes": observacoes,
        "metadata": metadata,
        "itens": itens,
    }

    return [{key: value for key, value in record.items() if value not in (None, "", [])}], "gerar_fatura"


def _parse_fatura_billing_item(data: dict[str, str]) -> dict[str, float | str] | None:
    descricao = data.get("Produto/Serviço")
    if not descricao or descricao == "TOTAL R$":
        return None

    quantidade = _to_float(data.get("Vidas Ativas")) or 1.0
    valor_unitario = _to_float(data.get("Valor por Vida R$"))
    valor_total = _to_float(data.get("Total R$"))

    if valor_total is not None and valor_unitario is not None and quantidade:
        calculado = quantidade * valor_unitario
        if abs(calculado - valor_total) > 0.01:
            quantidade = 1.0
            valor_unitario = valor_total

    if valor_unitario is None and valor_total is not None:
        valor_unitario = valor_total
        quantidade = 1.0

    item = {
        "descricao": descricao,
        "quantidade": quantidade,
        "valor_unitario": valor_unitario,
    }
    item = {key: value for key, value in item.items() if value not in (None, "")}
    if item.get("descricao") and item.get("valor_unitario") is not None:
        return item

    return None


def _parse_fatura_employee_row(data: dict[str, str]) -> dict[str, str | float | int] | None:
    nome = _clean_name(data.get("Nome"))
    if not nome:
        return None

    matricula = _to_int(data.get("Matricula"))

    funcionario = {
        "setor": _clean_name(data.get("Setor")),
        "nome": nome,
        "situacao": _clean_name(data.get("Situação")),
        "matricula": matricula,
        "valor_cobrar": _to_float(data.get("Vl.Cobrar R$")),
    }

    return {key: value for key, value in funcionario.items() if value not in (None, "")}


def _parse_fatura_exam_row(data: dict[str, str]) -> dict[str, str | float]:
    nome = _clean_name(data.get("Nome"))
    if not nome:
        return None

    exame = {
        "nome": nome,
        "quantidade": _to_float(data.get("Quantidade")) or 1.0,
        "valor_cobrar": _to_float(data.get("Valor Cobrar R$")) or 0.0,
    }

    return {key: value for key, value in exame.items() if value not in (None, "")}


def _build_fatura_attachment_summary(
    *,
    unidade: str | None,
    itens: list[dict],
    funcionarios: list[dict],
    exames: list[dict],
    numero_funcionarios: int | None,
) -> str | None:
    partes: list[str] = ["Detalhes importados do anexo da fatura."]

    if unidade:
        partes.append(f"Unidade: {unidade}.")

    if itens:
        itens_resumo = ", ".join(_format_billing_item_summary(item) for item in itens[:6])
        sufixo_itens = ""
        if len(itens) > 6:
            sufixo_itens = f" + {len(itens) - 6} outro(s)"
        partes.append(f"Itens faturados: {itens_resumo}{sufixo_itens}.")

    if funcionarios:
        total_funcionarios = numero_funcionarios or len(funcionarios)
        funcionarios_resumo = ", ".join(_format_employee_summary(funcionario) for funcionario in funcionarios[:8])
        sufixo_funcionarios = ""
        if total_funcionarios > 8:
            sufixo_funcionarios = f" + {total_funcionarios - 8} outro(s)"
        partes.append(
            f"Funcionários vinculados ({total_funcionarios}): {funcionarios_resumo}{sufixo_funcionarios}."
        )
    elif numero_funcionarios:
        partes.append(f"Funcionários vinculados: {numero_funcionarios}.")

    if exames:
        exames_resumo = ", ".join(_format_exam_summary(exame) for exame in exames[:10])
        sufixo_exames = ""
        if len(exames) > 10:
            sufixo_exames = f" + {len(exames) - 10} outro(s)"
        partes.append(f"Exames identificados: {exames_resumo}{sufixo_exames}.")

    summary = " ".join(parte for parte in partes if parte).strip()
    return summary or None


def _format_billing_item_summary(item: dict[str, str | float]) -> str:
    descricao = str(item.get("descricao") or "Item")
    quantidade = _to_float(item.get("quantidade"))
    valor_unitario = _to_float(item.get("valor_unitario"))

    if quantidade and quantidade != 1 and valor_unitario is not None:
        return f"{descricao} ({_format_number(quantidade)} x {_format_currency(valor_unitario)})"
    if valor_unitario is not None:
        return f"{descricao} ({_format_currency(valor_unitario)})"

    return descricao


def _format_employee_summary(funcionario: dict[str, str | float | int]) -> str:
    nome = str(funcionario.get("nome") or "Sem nome")
    detalhes: list[str] = []

    if funcionario.get("setor"):
        detalhes.append(str(funcionario["setor"]))
    if funcionario.get("matricula") is not None:
        detalhes.append(f"matrícula {funcionario['matricula']}")
    if funcionario.get("situacao"):
        detalhes.append(str(funcionario["situacao"]).lower())

    if detalhes:
        return f"{nome} ({', '.join(detalhes)})"

    return nome


def _format_exam_summary(exame: dict[str, str | float]) -> str:
    nome = str(exame.get("nome") or "Exame")
    quantidade = _to_float(exame.get("quantidade"))
    valor = _to_float(exame.get("valor_cobrar"))

    partes: list[str] = []
    if quantidade is not None:
        partes.append(f"{_format_number(quantidade)}x")
    if valor is not None:
        partes.append(_format_currency(valor))

    if partes:
        return f"{nome} ({', '.join(partes)})"

    return nome


def _format_currency(value: float) -> str:
    formatted = f"{value:,.2f}"
    formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {formatted}"


def _format_number(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return str(value).replace(".", ",")


def _clean_name(value: str | None) -> str | None:
    if value in (None, ""):
        return None

    cleaned = " ".join(str(value).strip().split())
    return cleaned or None


def _to_int(value: str | None) -> int | None:
    if value in (None, ""):
        return None

    try:
        return int(float(str(value).strip()))
    except ValueError:
        return None


def _render_sheet_rows(*, title: str, rows) -> list[str]:
    lines: list[str] = [f"[Aba] {title}"]
    current_headers: dict[int, str] | None = None

    for row in rows:
        values = [_stringify_cell(value) for value in row]
        non_empty = [(index, value) for index, value in enumerate(values) if value]

        if not non_empty:
            current_headers = None
            continue

        if len(non_empty) == 1:
            lines.append(non_empty[0][1])
            current_headers = None
            continue

        visible_values = [value for _, value in non_empty]
        if _looks_like_header_row(visible_values):
            current_headers = {index: value for index, value in non_empty}
            lines.append("Cabecalhos: " + " | ".join(visible_values))
            continue

        if current_headers:
            mapped = {
                current_headers.get(index, f"coluna_{index + 1}"): value
                for index, value in non_empty
                if value
            }
            if mapped:
                lines.append("Registro: " + json.dumps(mapped, ensure_ascii=False))
                continue

        lines.append(
            "Linha: "
            + " | ".join(f"coluna_{index + 1}={value}" for index, value in non_empty)
        )

    return lines


def _looks_like_header_row(values: list[str]) -> bool:
    if len(values) < 2:
        return False

    alpha_like = sum(1 for value in values if any(char.isalpha() for char in value))
    numeric_like = sum(1 for value in values if _looks_like_data_value(value))

    return alpha_like >= max(2, len(values) - 1) and numeric_like == 0


def _looks_like_data_value(value: str) -> bool:
    normalized = value.strip()
    if normalized == "":
        return False

    if normalized.isdigit():
        return True

    compact = normalized.replace(".", "").replace(",", "").replace("/", "").replace("-", "").replace(":", "").replace(" ", "")
    if compact.isdigit():
        return True

    if normalized.startswith("{") and normalized.endswith("}"):
        return False

    return False


def _stringify_cell(value) -> str:
    if value in (None, ""):
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)

    return str(value).strip()


def _normalize_row_keys(row: dict[str, str]) -> dict[str, str]:
    return {
        _normalize_header_name(key): value
        for key, value in row.items()
        if key
    }


def _normalize_header_name(value: str) -> str:
    text = str(value or "").strip().lower().lstrip("\ufeff")
    replacements = {
        "ã": "a",
        "á": "a",
        "à": "a",
        "â": "a",
        "é": "e",
        "ê": "e",
        "í": "i",
        "ó": "o",
        "ô": "o",
        "õ": "o",
        "ú": "u",
        "ç": "c",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)

    cleaned = []
    for char in text:
        cleaned.append(char if char.isalnum() else "_")

    normalized = "".join(cleaned)
    while "__" in normalized:
        normalized = normalized.replace("__", "_")
    return normalized.strip("_")


def _normalize_date_value(value: str | None) -> str | None:
    if not value:
        return None

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(value[:19], fmt).date().isoformat()
        except ValueError:
            continue

    return None


def _to_float(value: str | None) -> float | None:
    if value in (None, ""):
        return None

    text = str(value).strip()
    if text == "":
        return None

    allowed = "".join(char for char in text if char.isdigit() or char in {",", ".", "-"})
    if allowed == "":
        return None

    if "," in allowed and "." in allowed:
        allowed = allowed.replace(".", "").replace(",", ".")
    elif "," in allowed:
        allowed = allowed.replace(",", ".")

    try:
        return float(allowed)
    except ValueError:
        return None


def _truncate(text: str, max_chars: int) -> str:
    text = text.strip()
    if len(text) <= max_chars:
        return text
    return text[:max_chars].rstrip() + "\n...[conteudo truncado]"
